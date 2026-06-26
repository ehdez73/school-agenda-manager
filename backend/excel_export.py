"""
Excel export for timetables using openpyxl.
Preserves cell colours from Subject, SubjectGroup, FixedSlot and Config.day_colors.
Exports one sheet per course (by course_line) and one sheet per teacher (by name).
"""

import json
from io import BytesIO
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import (
    TimeSlotAssignment,
    Timeslot,
    Teacher,
    Config,
    FixedSlot,
    TeacherBusySlot,
    SupportAssignment,
    JointClass,
    TeacherFixedSlotLabel,
    CourseFixedSlotLabel,
    normalize_tutor_groups,
)
from .translations import t

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

COORDINATION_COLOR = "#e8f5e9"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
HOUR_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
CONFLICT_FONT = Font(bold=True, color="eb5252", size=10)

HEADER_FILL = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")

HEADER_FONT_WHITE = Font(bold=True, size=10, color="ffffff")
HEADER_FILL_BLACK = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
HOUR_FILL = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
COORD_FILL = PatternFill(
    start_color=COORDINATION_COLOR.lstrip("#"),
    end_color=COORDINATION_COLOR.lstrip("#"),
    fill_type="solid",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _hex_to_fill(hex_color):
    """Convert a hex colour string (e.g. ``"#dbeafe"``) to an openpyxl PatternFill
    or return *None* if the value is invalid/missing."""
    if not hex_color:
        return None
    color = hex_color.lstrip("#")
    if len(color) != 6:
        return None
    try:
        int(color, 16)
    except ValueError:
        return None
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _interleave_rows(sorted_hours, fixed_slots_sorted):
    """Yield (row_type, data) tuples interleaving solver hours and fixed slots
    by their 1-indexed ``position``."""
    solver_idx = 0
    fixed_idx = 0
    position = 1

    while solver_idx < len(sorted_hours) or fixed_idx < len(fixed_slots_sorted):
        next_fixed = (
            fixed_slots_sorted[fixed_idx]
            if fixed_idx < len(fixed_slots_sorted)
            else None
        )
        next_fixed_pos = next_fixed.position if next_fixed else None

        if next_fixed_pos is not None and next_fixed_pos == position:
            yield ("fixed", next_fixed)
            fixed_idx += 1
        elif solver_idx < len(sorted_hours):
            yield ("solver", sorted_hours[solver_idx])
            solver_idx += 1
        position += 1


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def _get_course_data(session, course_lines=None):
    """Build structured course-timetable data with labels and colours.

    Returns a dict keyed by course_line (e.g. ``"1ºA"``).  Each value contains
    ``cells`` (``{(hour, day): [(label, colour)]}``), ``sorted_hours``,
    ``weekdays``, ``hour_names``, etc.
    """
    # Tutors lookup
    tutors_dict = {}
    for teacher in session.query(Teacher).all():
        for tutor_group in normalize_tutor_groups(teacher.tutor_group):
            tutors_dict[tutor_group] = teacher.name

    # Timetable cells
    timetable = defaultdict(lambda: defaultdict(list))
    assignments = session.query(TimeSlotAssignment).all()
    for assignment in assignments:
        ts = assignment.timeslot
        course_line = f"{ts.course_id}{chr(ord('A') + ts.line)}"
        hour, weekday = ts.hour, ts.day
        subject_name = assignment.subject.name
        teacher_name = assignment.teacher.name

        group_color = None
        if ts.subject_group is not None:
            group_color = getattr(ts.subject_group, "color", None)
        subject_color = group_color or assignment.subject.color
        label = f"{subject_name} ({teacher_name})"
        timetable[course_line][(hour, weekday)].append((label, subject_color))

    # Support assignments
    support_assignments = session.query(SupportAssignment).all()
    for sa in support_assignments:
        course_line = f"{sa.course_id}{chr(ord('A') + sa.line)}"
        label = f"{sa.teacher.name} ({t('timetable.support_label')})"
        timetable[course_line][(sa.hour, sa.day)].append((label, sa.subject.color))

    # Filter by requested course_lines (empty list = none)
    if course_lines is not None:
        cl_set = set(course_lines)
        timetable = {k: v for k, v in timetable.items() if k in cl_set}

    # Config
    cfg = session.query(Config).first()
    cfg_dict = cfg.to_dict() if cfg else {}
    hour_names = cfg_dict.get("hour_names", []) if cfg_dict else []
    days_per_week = cfg_dict.get("days_per_week", 5) if cfg_dict else 5
    day_indices_from_cfg = cfg_dict.get("day_indices") if cfg_dict else None
    day_colors_raw = cfg_dict.get("day_colors") if cfg_dict else {}

    # Fixed slots (course type)
    course_fixed_slots = sorted(
        session.query(FixedSlot).filter_by(slot_type="course").all(),
        key=lambda fs: fs.position,
    )
    override_rows = session.query(CourseFixedSlotLabel).all()
    course_fixed_labels = {}
    for row in override_rows:
        course_fixed_labels.setdefault(row.course_line, {}).setdefault(
            row.fixed_slot_id, {}
        )[row.day] = row.label

    # Build per-course result dict
    result = {}
    for course_line, slots in timetable.items():
        sorted_hours = sorted({h for (h, _) in slots.keys()})
        day_indices = (
            day_indices_from_cfg
            if day_indices_from_cfg is not None
            else list(range(days_per_week))
        )
        weekdays = [t(f"day.{i}") for i in day_indices]
        tutor_name = tutors_dict.get(course_line)
        support_hours = sum(
            1
            for sa in support_assignments
            if f"{sa.course_id}{chr(ord('A') + sa.line)}" == course_line
        )

        result[course_line] = {
            "tutor_name": tutor_name,
            "support_hours": support_hours,
            "day_indices": day_indices,
            "weekdays": weekdays,
            "hour_names": hour_names,
            "sorted_hours": sorted_hours,
            "cells": dict(slots),
            "day_colors": day_colors_raw,
            "fixed_rows": course_fixed_slots,
            "fixed_labels": course_fixed_labels.get(course_line, {}),
        }

    return result


def _get_teacher_data(session, teacher_names=None):
    """Build structured teacher-timetable data with labels and colours.

    Returns a dict keyed by ``teacher.name``.  Each value contains ``cells``
    (``{(hour, day): [(label, colour, is_conflict)]}``) and metadata.
    """
    # JointClass lookup
    jc_lookup = defaultdict(list)
    for jc in session.query(JointClass).all():
        lines = json.loads(jc.lines) if isinstance(jc.lines, str) else jc.lines
        for line_letter in lines:
            jc_lookup[(jc.course_id, jc.subject_id, line_letter)].append(jc)

    # Teacher unavailable slots
    teacher_unavailable = {}
    for teacher in session.query(Teacher).all():
        prefs = {}
        if teacher.preferences:
            try:
                prefs = json.loads(teacher.preferences)
            except (ValueError, TypeError):
                prefs = {}
        unavailable_set = set()
        for day_str, day_prefs in prefs.items():
            if isinstance(day_prefs, dict) and "unavailable" in day_prefs:
                for h in day_prefs["unavailable"]:
                    unavailable_set.add((int(day_str), h))
        teacher_unavailable[teacher.name] = unavailable_set

    # First pass: collect raw items
    cell_items = defaultdict(list)
    assignments = session.query(TimeSlotAssignment).all()
    for assignment in assignments:
        ts = assignment.timeslot
        teacher_name = assignment.teacher.name
        subject_name = assignment.subject.name
        course_line = f"{ts.course_id}{chr(ord('A') + ts.line)}"
        line_letter = chr(ord("A") + ts.line)

        matched_jc = None
        for jc in jc_lookup.get((ts.course_id, assignment.subject_id, line_letter), []):
            if jc.teacher_id is None or jc.teacher_id == assignment.teacher_id:
                matched_jc = jc
                break

        group_color = None
        if ts.subject_group is not None:
            group_color = getattr(ts.subject_group, "color", None)
        subject_color = group_color or assignment.subject.color

        cell_items[(teacher_name, ts.hour, ts.day)].append({
            "jc": matched_jc,
            "subject_name": subject_name,
            "subject_id": assignment.subject.id,
            "course_line": course_line,
            "subject_color": subject_color,
        })

    # Second pass: build structured cells
    teacher_data = defaultdict(lambda: defaultdict(list))
    for (teacher_name, hour, day), items in cell_items.items():
        jc_groups = defaultdict(list)
        for item in items:
            key = id(item["jc"]) if item["jc"] else None
            jc_groups[key].append(item)

        is_conflict = (day, hour) in teacher_unavailable.get(teacher_name, set())

        for group_key, group in jc_groups.items():
            if group_key is not None and len(group) > 1:
                jc = group[0]["jc"]
                if jc and jc.name:
                    jc_display = jc.name
                else:
                    jc_display = f"{group[0]['subject_name']} ({t('timetable.joint_class_label')})"
                teacher_data[teacher_name][(hour, day)].append(
                    (group[0]["subject_id"], group[0]["subject_color"], is_conflict, False, "", jc_display)
                )
            else:
                for item in group:
                    teacher_data[teacher_name][(hour, day)].append(
                        (item["subject_id"], item["subject_color"], is_conflict, False, item["course_line"], item["subject_name"])
                    )

    # Support
    for sa in session.query(SupportAssignment).all():
        teacher_name = sa.teacher.name
        subject_id = sa.subject.id
        subject_name = sa.subject.name
        is_conflict = (sa.day, sa.hour) in teacher_unavailable.get(teacher_name, set())
        support_course_line = f"{sa.course_id}{chr(ord('A') + sa.line)}"
        teacher_data[teacher_name][(sa.hour, sa.day)].append(
            (subject_id, sa.subject.color, is_conflict, True, support_course_line, subject_name)
        )

    # Busy slots (coordination)
    for slot in session.query(TeacherBusySlot).all():
        teacher_data[slot.teacher.name][(slot.hour, slot.day)].append(
            (t("timetable.coordination_label"), COORDINATION_COLOR, False, False, "", "")
        )

    # Fill gaps
    cfg = session.query(Config).first()
    num_days = cfg.days_per_week if cfg else 5
    num_hours = cfg.classes_per_day if cfg else 5

    for teacher_name, slots in list(teacher_data.items()):
        unavailable = teacher_unavailable.get(teacher_name, set())
        for d in range(num_days):
            for h in range(num_hours):
                key = (h, d)
                if key not in slots:
                    if (d, h) in unavailable:
                        slots[key].append(("✕", None, True, False, "", ""))
                    else:
                        slots[key].append(("", None, False, False, "", ""))

    # Filter by teacher_names (empty list = none)
    if teacher_names is not None:
        tn_set = set(teacher_names)
        teacher_data = {k: v for k, v in teacher_data.items() if k in tn_set}

    # Teacher hours (from teacher_utils)
    from .teacher_utils import compute_teacher_hours

    hours_data = compute_teacher_hours(session)
    teacher_info = {}
    for tid, hd in hours_data.items():
        teacher_info[hd["name"]] = hd

    cfg_dict = cfg.to_dict() if cfg else {}
    hour_names = cfg_dict.get("hour_names", []) if cfg_dict else []
    days_per_week = cfg_dict.get("days_per_week", 5) if cfg_dict else 5
    day_indices_from_cfg = cfg_dict.get("day_indices") if cfg_dict else None
    day_colors_raw = cfg_dict.get("day_colors") if cfg_dict else {}

    # Fixed slots (teacher type)
    teacher_fixed_slots = sorted(
        session.query(FixedSlot).filter_by(slot_type="teacher").all(),
        key=lambda fs: fs.position,
    )

    # Teacher fixed slot labels
    override_rows = session.query(TeacherFixedSlotLabel).all()
    teacher_fixed_labels = {}
    for row in override_rows:
        teacher_fixed_labels.setdefault(row.teacher.name, {}).setdefault(
            row.fixed_slot_id, {}
        )[row.day] = row.label

    result = {}
    for teacher_name, slots in teacher_data.items():
        sorted_hours = sorted({h for (h, _) in slots.keys()})
        day_indices = (
            day_indices_from_cfg
            if day_indices_from_cfg is not None
            else list(range(days_per_week))
        )
        weekdays = [t(f"day.{i}") for i in day_indices]
        info = teacher_info.get(teacher_name, {})

        result[teacher_name] = {
            "max_hours": info.get("max_hours_week", 0),
            "assigned_hours": info.get("assigned_hours", 0),
            "coord_hours": info.get("coordination_hours", 0),
            "support_hours": info.get("support_hours", 0),
            "day_indices": day_indices,
            "weekdays": weekdays,
            "hour_names": hour_names,
            "sorted_hours": sorted_hours,
            "cells": dict(slots),
            "day_colors": day_colors_raw,
            "fixed_rows": teacher_fixed_slots,
            "fixed_labels": teacher_fixed_labels.get(teacher_name, {}),
        }

    return result


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_course_sheet(ws, course_line, info):
    """Write a single course timetable into *ws*."""
    day_indices = info["day_indices"]
    weekdays = info["weekdays"]
    hour_names = info["hour_names"]
    sorted_hours = info["sorted_hours"]
    cells = info["cells"]
    tutor_name = info["tutor_name"]
    support_hours = info["support_hours"]
    fixed_rows = info.get("fixed_rows", [])
    fixed_labels = info.get("fixed_labels", {})
    day_colors = info.get("day_colors", {})

    num_days = len(day_indices)

    # --- Row 1: Title ---
    title_parts = [f"{t('timetable.course_label')}: {course_line}"]
    if tutor_name:
        title_parts.append(f"\u2014 {t('timetable.group_tutor')}: {tutor_name}")
    if support_hours > 0:
        title_parts.append(
            f"({support_hours}h {t('timetable.support_label_short')})"
        )
    title = " ".join(title_parts)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=num_days + 1)
    c = ws.cell(row=1, column=2, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 2: Headers ---
    c = ws.cell(row=2, column=1, value=t("timetable.hour_header"))
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL_BLACK
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, day_idx in enumerate(day_indices, start=2):
        day_name = weekdays[col_idx - 2]
        c = ws.cell(row=2, column=col_idx, value=day_name)
        c.font = HEADER_FONT_WHITE
        c.border = THIN_BORDER
        c.fill = HEADER_FILL_BLACK
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 40

    # --- Data rows ---
    data_row = 3
    for row_type, data in _interleave_rows(sorted_hours, fixed_rows):
        if row_type == "fixed":
            fs = data
            slot_overrides = fixed_labels.get(fs.id, {})
            hour_label = fs.time_range

            c = ws.cell(row=data_row, column=1, value=hour_label)
            c.font = HOUR_FONT
            c.fill = HOUR_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

            for day_index in range(num_days):
                override = slot_overrides.get(day_index)
                label = override if override is not None else fs.label
                c = ws.cell(row=data_row, column=day_index + 2, value=label)
                c.border = THIN_BORDER
                c.font = NORMAL_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
                if fs.color:
                    fill = _hex_to_fill(fs.color)
                    if fill:
                        c.fill = fill
        else:
            hour = data
            hour_label = (
                hour_names[hour]
                if hour < len(hour_names)
                else t("hours.label").format(n=hour + 1)
            )
            c = ws.cell(row=data_row, column=1, value=hour_label)
            c.font = HOUR_FONT
            c.fill = HOUR_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

            max_lines = 1
            for day_index in range(num_days):
                assignments = cells.get((hour, day_index), [])
                if assignments:
                    lines = []
                    colours = set()
                    for label, colour in assignments:
                        lines.append(label)
                        if colour:
                            colours.add(colour)
                    cell_text = "\n".join(lines)
                    c = ws.cell(
                        row=data_row, column=day_index + 2, value=cell_text
                    )
                    c.border = THIN_BORDER
                    c.font = NORMAL_FONT
                    c.alignment = Alignment(
                        wrap_text=True, vertical="center", horizontal="center"
                    )

                    # If all items share the same colour apply it to the cell
                    if len(colours) == 1:
                        fill = _hex_to_fill(next(iter(colours)))
                        if fill:
                            c.fill = fill
                    max_lines = max(max_lines, len(lines))
                else:
                    c = ws.cell(
                        row=data_row, column=day_index + 2, value=""
                    )
                    c.border = THIN_BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[data_row].height = 80

        data_row += 1

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, num_days + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 32

    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Teacher grid sheet (single sheet for all teachers, day-by-day sections)
# ---------------------------------------------------------------------------

def _write_teacher_grid_sheet(ws, teacher_data, day_indices, weekdays,
                              hour_names, day_colors):
    """Write all teacher timetables into one sheet as a grid with one section
    per day, matching the ``teacher_staff`` front-end tab layout."""
    teacher_names = sorted(teacher_data.keys())
    if not teacher_names:
        return

    num_hours = len(hour_names)
    num_days = len(day_indices)
    num_cols = len(teacher_names) + 1

    # --- Row 1: Title ---
    ws.merge_cells(
        start_row=1, start_column=2, end_row=1, end_column=num_cols
    )
    c = ws.cell(row=1, column=2, value=t("timetable.tab_teacher_staff"))
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 2: Column headers (shown once, rotated vertically up) ---
    header_alignment = Alignment(text_rotation=90, vertical='center', horizontal='center')
    c = ws.cell(row=2, column=1, value=t("timetable.hour_header"))
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL_BLACK
    c.border = THIN_BORDER
    c.alignment = header_alignment
    for ci, t_name in enumerate(teacher_names, start=2):
        c = ws.cell(row=2, column=ci, value=t_name)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL_BLACK
        c.border = THIN_BORDER
        c.alignment = header_alignment
    max_name_len = max((len(n) for n in teacher_names), default=0)
    ws.row_dimensions[2].height = max_name_len * 9

    row = 4  # skip title row + header row + blank separator
    for di, day_idx in enumerate(day_indices):
        day_name = weekdays[di]

        # Day section header (merged, with day colour)
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=num_cols
        )
        c = ws.cell(row=row, column=1, value=day_name)
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="left", vertical="center")
        day_color = (
            day_colors.get(str(day_idx))
            if isinstance(day_colors, dict)
            else None
        )
        if day_color:
            fill = _hex_to_fill(day_color)
            if fill:
                c.fill = fill
        row += 1

        # Data rows
        for hour in range(num_hours):
            hour_label = (
                hour_names[hour]
                if hour < len(hour_names)
                else t("hours.label").format(n=hour + 1)
            )
            c = ws.cell(row=row, column=1, value=hour_label)
            c.font = HOUR_FONT
            c.fill = HOUR_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

            max_lines = 1
            for ci, t_name in enumerate(teacher_names, start=2):
                items = teacher_data[t_name]["cells"].get(
                    (hour, day_idx), []
                )
                if items:
                    lines = []
                    is_unavailable = False
                    is_support = False
                    is_empty_gap = False
                    any_conflict = False
                    for label, colour, conflict, support, course_line, _ in items:
                        if course_line:
                            lines.append(f"{course_line}-{label}")
                        else:
                            lines.append(label)
                        if conflict:
                            any_conflict = True
                        if label == "\u2715":
                            is_unavailable = True
                        if support:
                            is_support = True
                    is_empty_gap = all(label == "" for label, _, _, _, _, _ in items)

                    cell_text = "\n".join(lines)
                    cell_text = cell_text.strip() or ""
                    c = ws.cell(row=row, column=ci, value=cell_text)
                    c.border = THIN_BORDER
                    c.font = CONFLICT_FONT if any_conflict else NORMAL_FONT
                    c.alignment = Alignment(
                        wrap_text=True, vertical="center", horizontal="center"
                    )

                    if is_unavailable:
                        c.fill = _hex_to_fill("#eb5252")
                    elif is_support or is_empty_gap:
                        pass
                    elif day_color:
                        fill = _hex_to_fill(day_color)
                        if fill:
                            c.fill = fill
                    max_lines = max(max_lines, len(lines))
                else:
                    c = ws.cell(row=row, column=ci, value="")
                    c.border = THIN_BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row].height = max(15, max_lines * 15)
            row += 1

        row += 1  # blank separator between days

    # Auto-fit column widths
    max_hour_len = max((len(str(h)) for h in hour_names), default=0)
    ws.column_dimensions["A"].width = max(max_hour_len + 2, 8)
    for ci, t_name in enumerate(teacher_names, start=2):
        col_letter = get_column_letter(ci)
        max_len = 0
        for (hour, day_idx), items in teacher_data[t_name]["cells"].items():
            for label, colour, conflict, support, course_line, _ in items:
                if course_line:
                    display = f"{course_line}-{label}"
                else:
                    display = label
                max_len = max(max_len, len(display))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 8)

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Individual teacher sheet
# ---------------------------------------------------------------------------

def _write_teacher_sheet(ws, teacher_name, info):
    """Write a single teacher timetable into *ws*."""
    day_indices = info["day_indices"]
    weekdays = info["weekdays"]
    hour_names = info["hour_names"]
    sorted_hours = info["sorted_hours"]
    cells = info["cells"]
    fixed_rows = info.get("fixed_rows", [])
    fixed_labels = info.get("fixed_labels", {})
    day_colors = info.get("day_colors", {})

    num_days = len(day_indices)

    # --- Row 1: Title ---
    title_parts = [f"{t('timetable.teacher_label')}: {teacher_name}"]
    summary_parts = []
    max_hours = info.get("max_hours", 0)
    assigned = info.get("assigned_hours", 0)
    coord = info.get("coord_hours", 0)
    support = info.get("support_hours", 0)
    if max_hours:
        summary_parts.append(f"{t('timetable.max_label')}: {max_hours}h")
    if assigned:
        summary_parts.append(f"{t('timetable.assigned_label')}: {assigned}h")
    if coord:
        summary_parts.append(f"{t('timetable.coordination_label_short')}: {coord}h")
    if support:
        summary_parts.append(f"{t('timetable.support_label')}: {support}h")
    if summary_parts:
        title_parts.append("\u2014 " + ", ".join(summary_parts))
    title = " ".join(title_parts)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=num_days + 1)
    c = ws.cell(row=1, column=2, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 2: Headers ---
    c = ws.cell(row=2, column=1, value=t("timetable.hour_header"))
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL_BLACK
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, day_idx in enumerate(day_indices, start=2):
        day_name = weekdays[col_idx - 2]
        c = ws.cell(row=2, column=col_idx, value=day_name)
        c.font = HEADER_FONT_WHITE
        c.border = THIN_BORDER
        c.fill = HEADER_FILL_BLACK
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 40

    # --- Data rows ---
    data_row = 3
    for row_type, data in _interleave_rows(sorted_hours, fixed_rows):
        if row_type == "fixed":
            fs = data
            slot_overrides = fixed_labels.get(fs.id, {})
            hour_label = fs.time_range

            c = ws.cell(row=data_row, column=1, value=hour_label)
            c.font = HOUR_FONT
            c.fill = HOUR_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

            for day_index in range(num_days):
                override = slot_overrides.get(day_index)
                label = override if override is not None else fs.label
                c = ws.cell(row=data_row, column=day_index + 2, value=label)
                c.border = THIN_BORDER
                c.font = NORMAL_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
                if fs.color:
                    fill = _hex_to_fill(fs.color)
                    if fill:
                        c.fill = fill
        else:
            hour = data
            hour_label = (
                hour_names[hour]
                if hour < len(hour_names)
                else t("hours.label").format(n=hour + 1)
            )
            c = ws.cell(row=data_row, column=1, value=hour_label)
            c.font = HOUR_FONT
            c.fill = HOUR_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

            max_lines = 1
            for day_index in range(num_days):
                items = cells.get((hour, day_index), [])
                if items:
                    lines = []
                    colours = set()
                    any_conflict = False
                    is_unavailable = False
                    for label, colour, conflict, support, course_line, subject_name in items:
                        if label == "\u2715":
                            lines.append("\u2715")
                            is_unavailable = True
                            any_conflict = True
                        elif support and subject_name:
                            display = f"{course_line}: {subject_name} ({t('timetable.support_label')})"
                            lines.append(display)
                            if colour:
                                colours.add(colour)
                            if conflict:
                                any_conflict = True
                        elif isinstance(label, str) and label:
                            if course_line:
                                display = f"{course_line}: {subject_name}"
                            else:
                                display = subject_name
                            lines.append(display)
                            if colour:
                                colours.add(colour)
                        elif subject_name:
                            if course_line:
                                display = f"{course_line}: {subject_name}"
                            else:
                                display = subject_name
                            lines.append(display)
                            if colour:
                                colours.add(colour)
                            if conflict:
                                any_conflict = True
                        else:
                            lines.append(str(label))
                    cell_text = "\n".join(lines)
                    c = ws.cell(
                        row=data_row, column=day_index + 2, value=cell_text
                    )
                    c.border = THIN_BORDER
                    if is_unavailable:
                        c.font = CONFLICT_FONT
                        fill = _hex_to_fill("#eb5252")
                        if fill:
                            c.fill = fill
                    elif any_conflict:
                        c.font = CONFLICT_FONT
                        if len(colours) == 1:
                            fill = _hex_to_fill(next(iter(colours)))
                            if fill:
                                c.fill = fill
                    else:
                        c.font = NORMAL_FONT
                        if len(colours) == 1:
                            fill = _hex_to_fill(next(iter(colours)))
                            if fill:
                                c.fill = fill
                    c.alignment = Alignment(
                        wrap_text=True, vertical="center", horizontal="center"
                    )
                    max_lines = max(max_lines, len(lines))
                else:
                    c = ws.cell(
                        row=data_row, column=day_index + 2, value=""
                    )
                    c.border = THIN_BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[data_row].height = 80

        data_row += 1

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, num_days + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 32

    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_excel_timetable(session, course_lines=None, teacher_names=None,
                             teacher_grouped=True):
    """Build an ``.xlsx`` workbook with one sheet per course and either a
    single ``Profesorado`` grid sheet or individual teacher sheets.

    Parameters
    ----------
    session : sqlalchemy.orm.Session
        Database session.
    course_lines : list[str] | None
        Course lines to include (e.g. ``["1\u00baA", "2\u00baB"]``).
        *None* means all.
    teacher_names : list[str] | None
        Teacher names to include.  *None* means all.
    teacher_grouped : bool
        If *True* (default), create a single grid sheet with all teachers.
        If *False*, create one sheet per teacher individually.

    Returns
    -------
    BytesIO
        In-memory Excel file ready to be served as a download.
    """
    course_data = _get_course_data(session, course_lines)
    teacher_data = _get_teacher_data(session, teacher_names)

    wb = Workbook()
    wb.remove(wb.active)

    # Course sheets (one per course)
    for course_line in sorted(course_data):
        sheet_name = str(course_line)[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_course_sheet(ws, course_line, course_data[course_line])

    # Teacher sheets
    if teacher_data:
        if teacher_grouped:
            # Single grid sheet (Profesorado tab)
            ws = wb.create_sheet(
                title=(t("timetable.tab_teacher_staff") or "Profesorado")[:31]
            )
            sample = next(iter(teacher_data.values()))
            _write_teacher_grid_sheet(
                ws,
                teacher_data,
                day_indices=sample["day_indices"],
                weekdays=sample["weekdays"],
                hour_names=sample["hour_names"],
                day_colors=sample.get("day_colors", {}),
            )
        else:
            # Individual teacher sheets (General tab)
            for teacher_name in sorted(teacher_data):
                sheet_name = str(teacher_name)[:31]
                ws = wb.create_sheet(title=sheet_name)
                _write_teacher_sheet(ws, teacher_name, teacher_data[teacher_name])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
